import openpyxl
import app

wb = openpyxl.load_workbook('sample_gemini.xlsx')
ws = wb['Sheet9']

val_str = '=VLOOKUP(CONCATENATE(A3,"|",B3),$O$3:$P$16,2,FALSE)'

import re
from openpyxl.utils.cell import coordinate_to_tuple

def debug_vlookup(val_str, ws):
    pattern = r'=VLOOKUP\((?:CONCATENATE\(([A-Z]+)([0-9]+)\s*,\s*\"\|\"\s*,\s*([A-Z]+)([0-9]+)\)|([A-Z]+)([0-9]+)\s*\&\s*\"\|\"\s*\&\s*([A-Z]+)([0-9]+))\s*,\s*\$?([A-Z]+)\$?([0-9]+):\$?([A-Z]+)\$?([0-9]+)\s*,\s*2\s*,\s*FALSE\)'
    match = re.search(pattern, val_str)
    
    if not match: 
        print("NO MATCH")
        return val_str
    groups = match.groups()
    
    if groups[0] is not None:
        col1, row1, col2, row2 = groups[0], groups[1], groups[2], groups[3]
    else:
        col1, row1, col2, row2 = groups[4], groups[5], groups[6], groups[7]
        
    start_col, start_row, end_col, end_row = groups[8], groups[9], groups[10], groups[11]
    
    print(f"Parsed: {col1}{row1} & {col2}{row2}, Range: {start_col}{start_row}:{end_col}{end_row}")
    # 1. Get search key
    part1 = ws[f"{col1}{row1}"].value
    part2 = ws[f"{col2}{row2}"].value
    
    print(f"Part1: {part1}, Part2: {part2}")
    if part1 is None: part1 = ""
    if part2 is None: part2 = ""
    search_key = f"{str(part1).strip()}|{str(part2).strip()}"
    print(f"Search Key: {search_key}")
    
    start_r, start_c = coordinate_to_tuple(f"{start_col}{start_row}")
    end_r, end_c = coordinate_to_tuple(f"{end_col}{end_row}")
    
    for r in range(start_r, end_r + 1):
        lookup_val = ws.cell(row=r, column=start_c).value
        # print(f"Row {r}: '{lookup_val}'")
        if lookup_val is not None and str(lookup_val).strip() == search_key:
            res = ws.cell(row=r, column=start_c + 1).value
            print(f"Found! returning {res}")
            return res if res is not None else ""
            
    print("Not found")
    return ""

debug_vlookup(val_str, ws)
